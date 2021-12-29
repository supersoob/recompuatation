import openpyxl
import os
import sys



cnt = 0
layer_cnt = 0
#layers = [ 0 for _ in range(31)]

if __name__ == "__main__":
    filename = sys.argv[1]
    f = open(filename, "r")
    lines = f.readlines()

    remove_idx = 0
    for i, l in enumerate(lines):
        if "Epoch" in l:
            if " 5/" in l:
                remove_idx = i+1
                break
    
    del lines[0:remove_idx]

    #print(lines)
        
    num_layer = 0
    for i, l in enumerate(lines):
        if "Epoch" in l:
            break
        num_layer = num_layer + 1


    
    layers = [0 for _ in range(num_layer)]
                    
    #print(lines)

    for l in lines:

       if "Epoch" in l:
            cnt = cnt+1
            layer_cnt = 0
       else:
            if "Finished"  in l:
                continue
            layers[layer_cnt] += float(l.strip())
            layer_cnt = layer_cnt + 1
#    print (cnt)

    for tm in layers:
        new_tm = (tm / cnt)
        print(new_tm)
        #sheet.append([new_tm])

    #wb.save("./Time_layers_vgg_imagenet.xlsx")
"""
    try:
    	wb = openpyxl.load_workbook(filename="Time_layers_vgg_imagenet.xlsx")
    except:
    	wb = openpyxl.Workbook()

    sheet = wb.create_sheet(title=f"batch_{sys.argv[1]}")
"""
