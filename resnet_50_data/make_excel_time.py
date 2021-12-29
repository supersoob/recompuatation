import openpyxl
import sys

f = open(f"batch_{sys.argv[1]}_time.txt","r")

fline = f.readlines()

residual = []

for i in range(17):
    residual.append(0.0)

#print(residual)

cnt = 0

for l in fline:
    if "forward" in l:
        cnt = cnt + 1

    elif "residual" in l or "residaul" in l:
        print(l)
        l = l.split(" : ")
        idx = int(l[0].split(" ")[1])
        residual[idx] += float(l[1])


print(cnt)

print ("AVG Time of Residual\n")

s = 0


try:
    wb = openpyxl.load_workbook("Time_residual.xlsx")
except:
    wb = openpyxl.Workbook()
    
sheet = wb.create_sheet()
sheet.title = f"batch {sys.argv[1]}"
print("cnt : ", cnt)

#print(residual)

for i, res in enumerate(residual):
    avgTime = res / cnt
    avgTime = avgTime * 1000 # make s to ms
    sheet.append([i, avgTime])

    print(avgTime)

wb.save("Time_residual.xlsx")
wb.close()
