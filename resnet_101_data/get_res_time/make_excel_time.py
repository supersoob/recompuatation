import openpyxl
import sys

f = open(f"time_batch_{sys.argv[1]}.txt","r")

fline = f.readlines()

residual = []

for i in range(34):
    residual.append(0.0)

#print(residual)

cnt = 0

for l in fline:
    if "residual" in l:
        l = l.split(" : ")
        #print(l[0].split(" ")[1])

        idx = int(l[0].split(" ")[1])
        residual[idx] += float(l[1])

    if "forward" in l:
        #step = l.split("]")[0].split("[")[1]
        #cnt = int(step)
        cnt = cnt + 1

#cnt = cnt + 1
print(cnt)

print ("AVG Time of Residual\n")

s = 0


try:
    wb = openpyxl.load_workbook("Time_residual.xlsx")
except:
    wb = openpyxl.Workbook()
    
sheet = wb.create_sheet()
sheet.title = f"batch {sys.argv[1]}"
print("cnt : ", cnt)

for i, res in enumerate(residual):
    avgTime = res / cnt
    avgTime = avgTime * 1000 # make s to ms
    sheet.append([i, avgTime])

    print(avgTime)

wb.save("Time_residual.xlsx")
wb.close()
