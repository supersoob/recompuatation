import openpyxl
import sys

#@profile
def main():
    f = open(f"time_inside_batch_{sys.argv[1]}.txt","r")

    fline = f.readlines()

    residual = []
    cnt = 0

    

    for i in range(33):
        residual.append([0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0])

    print(residual)


    cnt = 0

    for l in fline:
        
        if "time : " in l:
            #print(l.split(" : ")[1])
            value = float(l.split(" : ")[1])
            
            #idx = 0

            if "conv1" in l:
                idx = 0
                residual[cnt] [idx] += value
            if "bn1" in l:
                idx = 1
                residual[cnt] [idx] += value
            if "relu1" in l:
                idx = 2
                residual[cnt] [idx] += value
            if "conv2" in l:
                idx = 3
                residual[cnt] [idx] += value
            if "bn2" in l:
                idx = 4
                residual[cnt] [idx] += value
            if "relu2" in l:
                idx = 5
                residual[cnt] [idx] += value
            if "conv3" in l:
                idx = 6
                residual[cnt] [idx] += value
            if "bn3" in l:
                idx = 7
                residual[cnt] [idx] += value
            if "shortcut" in l:
                idx = 8
                residual[cnt] [idx] += value
            if "relu " in l:
                idx = 9
                residual[cnt] [idx] += value


        if "residual" in l:
            cnt = cnt + 1

        if "forward" in l:
            step = int(l.split("]")[0].split("[")[1])
            cnt = 0
            


    step = step + 1

    print(residual)

    print ("AVG Time of Residual\n")


    try:
        wb = openpyxl.load_workbook("Time_residual.xlsx")
    except:
        wb = openpyxl.Workbook()
        
    sheet = wb.create_sheet()
    sheet.title = f"batch {sys.argv[1]}"

    for i, res in enumerate(residual):
        for j, ins in enumerate(res):
            avgTime = ins / step
            avgTime = avgTime * 1000 # make s to ms
            sheet.append(["residual "+str(i+1), j, avgTime])
        sheet.append([" "])

    wb.save("Time_inside_residual.xlsx")
    wb.close()


if __name__ == "__main__":
    main()
